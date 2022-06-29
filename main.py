from tkinter import *
import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import date,datetime


tela= Tk()

background = '#595958'
frameBorda= '#2C4001'
corFundoFrame= '#8AA626'
corEscura = '#A2BF63'
corClara = '#E5F2B3'
corTitulo = 'black'
corFonte = '#2C4001'

class Oppitions_txt():
    def ler_stk_tot (self):
        with open("data.txt", "r") as config:
            return config.read()
    def modificar(self,replace):
        with open("data.txt", "w") as config:
            return config.write(replace)

class Dados(Oppitions_txt):
    def dados_url(self):

        self.jogo = self.soup.find("span", attrs={"class": "calc_event"}).text

        self.info1 = self.soup.find("tr", attrs={"data-number": "0"})
        self.nome1 = self.info1.find("td", attrs={"class": "booker_c"}).text
        self.odd1 = self.info1.find("input", attrs={"class": "koefficient form-control inline-input w-number"})

        self.info2 = self.soup.find("tr", attrs={"data-number": "1"})
        self.nome2 = self.info2.find("td", attrs={"class": "booker_c"}).text
        self.odd2 = self.info2.find("input", attrs={"class": "koefficient form-control inline-input w-number"})

        self.info3 = self.soup.find("tr", attrs={"data-number": "2"})
        if self.info3 != None:
            self.nome3 = self.info3.find("td", attrs={"class": "booker_c"}).text
            self.odd3 = self.info3.find("input", attrs={"class": "koefficient form-control inline-input w-number"})
            self.entrada3 = self.info3.find("td", attrs= {"coeff_c coeff_name"}).text
            self.nome3.encode("utf8")
            self.odd3 = float(self.odd2['value'])
        else:
            self.odd3 = 0

        self.nome2.encode("utf8")
        self.nome1.encode("utf8")
        self.odd1 = float(self.odd1['value'])
        self.odd2 = float(self.odd2['value'])
        stktot = float(self.ler_stk_tot())

        odds = [self.odd1,self.odd2,self.odd3]
        ct = 0
        stks = []
        self.stkC = []


        for odd in odds:
            a = 1/odd
            stks.append(a)
            ct += a

        for stk in stks:

            b = (stk/ct)*stktot
            self.stkC.append(round(b,2))

        dados = [self.jogo,self.nome1,self.odd1,self.stkC[0],self.nome2,self.odd2,self.stkC[1]]
        if self.info3 != None: 
            dados.append(self.nome3)
            dados.append(self.odd3)
            dados.append(self.stkC[2])

        return dados

    def dados_tempo(self):
        self.dataAtual = date.today()
        self.dataPlanilha = self.dataAtual.strftime('%d/%m/%Y')
        self.horaAtual = datetime.now()
        self.horaPlanilha =self.horaAtual.strftime('%H:%M')
        
    

class Funcs (Dados,Oppitions_txt):

    def apagar(self):
        self.entry_url.delete(0,END)
        self.entry_odd_c1.delete(0, END)
        self.entry_odd_c2.delete(0, END)
        self.entry_stk_c1.delete(0, END)
        self.entry_stk_c2.delete(0, END)
    
   

    def up_url(self):
        self.url = self.entry_url.get()
        self.conteudo_url = requests.get(self.url)

        if self.conteudo_url.status_code == 200:

            self.soup =BeautifulSoup(self.conteudo_url.content, 'html.parser')
            self.soup.encode= "utf-8"
            self.chama_labels(self.dados_url())
                
        else: 
            print('Erro ao buscar a URl !')

    def alterar(self):
        self.odd_1 = self.entry_odd_c1.get()
        self.odd_2 = self.entry_odd_c2.get()
        self.stk_1 = self.entry_stk_c1.get()
        self.stk_2 = self.entry_stk_c2.get()
        self.chama_labels(o1=self.odd_1,o2=self.odd_2,s1=self.stk_1,s2=self.stk_2)

    def up_stk(self):
        self.stk = self.entry_stk_tot.get()
        self.stk_tot = self.modificar(self.stk)
        self.lb_stk_p = Label(self.frame_stake, text=(f"Stake Total = R${self.stk}"),bg=corFundoFrame)
        self.lb_stk_p.place(relx=0.1, rely=0.1,relwidth=0.8, relheight=0.2)
        self.estilos()

    def apagar_labels(self):
        for widget in Application.casa_1:
            widget.destroy()
    

    def salvar(self):
        caminho = "SureBet.xlsx"
        planilha = openpyxl.load_workbook(caminho)
        pagina = planilha['Planilha de Apostas']

        #acha a proxima linha vazia
        vazia = 17
        for row in pagina.iter_rows(min_row=17):
            if(row[5].value == None):
                break
            vazia += 3

        #Joga dados na planilha
        pagina.cell(row=vazia, column=4).value = self.dataPlanilha
        pagina.cell(row=vazia, column=5).value = self.nome1
        pagina.cell(row=vazia+1, column=5).value = self.nome2
        pagina.cell(row=vazia, column=6).value = self.horaPlanilha
        pagina.cell(row=vazia, column=7).value = self.jogo
        #Celulas de Stake
        if(self.entry_stk_c1 | self.entry_stk_c2 == ""):
            pagina.cell(row=vazia, column=9).value = float(self.s1)
            pagina.cell(row=vazia+1, column=9).value = float(self.s2)
        else:
            pagina.cell(row=vazia, column=9).value = float(self.entry_stk_c1.get())
            pagina.cell(row=vazia+1, column=9).value = float(self.ebtry_stk_c2.get())

        #Celulas de ODD
        if(self.entry_odd_c1 | self.entry_odd_c2 == ""):
            pagina.cell(row=vazia, column=10).value = float(self.odd1['value'])
            pagina.cell(row=vazia+1, column=10).value = float(self.odd2['value'])
        else:
            pagina.cell(row=vazia, column=10).value = float(self.entry_odd_c1.get())
            pagina.cell(row=vazia+1, column=10).value = float(self.entry_odd_c1.get())

        #salva
        planilha.save(caminho)

class Application (Funcs,Oppitions_txt):
    def __init__(self):
        self.tela = tela
        self.config()
        self.stk_tot = self.ler_stk_tot()
        self.frames()
        self.chama_widgets()
        tela.mainloop()

    def config(self):
        self.tela
        self.tela.title("Automação de Planilha SureBet")
        self.tela.configure(background=background)
        self.tela.geometry("700x500+788+311")
        self.tela.maxsize(width="900", height="700")
        self.tela.minsize(width="500", height="300")
    
    def frames (self):
        self.frame_url = Frame(self.tela,bd=4, bg=corFundoFrame, highlightbackground=frameBorda, highlightthickness=3)
        self.frame_url.place(relx=0.05, rely= 0.03, relwidth=0.55,relheight=0.3)

        self.frame_stake = Frame(self.tela,bd=4, bg=corFundoFrame, highlightbackground=frameBorda, highlightthickness=3)
        self.frame_stake.place(relx=0.65,rely=0.03,relwidth=0.3,relheight=0.3)

        self.frame_valores = Frame(self.tela,bd=4, bg=corFundoFrame, highlightbackground=frameBorda, highlightthickness=3)
        self.frame_valores.place(relx=0.05, rely=0.4,relwidth=0.9, relheight=0.55,)
    
    def chama_widgets(self):
        self.widgets_frame_stake()
        self.widgets_frame_url()
    
    def chama_labels(self,lista):
        self.label_nome(lista[0])
        self.casa_1(lista[1],lista[2],lista[3])
        self.casa_2(lista[4],lista[5],lista[6])
        if len(lista) > 7:
            self.casa_3(lista[7],lista[8],lista[9])
    
    def chama_tabela(self):
        self.estilos()
        self.widgets_frame_valores()

  
    def widgets_frame_stake(self):
        self.lb_stk_p = Label(self.frame_stake, bg=corFundoFrame,fg=corTitulo, font=('verdana',8, 'bold'), text=(f"Stake Total = R${self.stk_tot}"))
        self.lb_stk_p.place(relx=0.1, rely=0.1,relwidth=0.8, relheight=0.2)

        self.entry_stk_tot = Entry(self.frame_stake,bg=corClara,font=('verdana',8, 'bold'),bd=0)
        self.entry_stk_tot.place(relx=0.2,rely=0.35,relwidth=0.6,relheight=0.2)

        self.bt_stk_p = Button(self.frame_stake,text="Refinir",command=self.up_stk,bg=corClara,bd=2)
        self.bt_stk_p.place(relx=0.2,rely=0.65,relwidth=0.6,relheight=0.2)

    def estilos (self):
        self.frame_estilo1 = Frame(self.frame_valores, bd=0, highlightthickness=0,bg=corEscura)
        self.frame_estilo1.place(relx=-0.007,rely=0.22,relwidth=1.014,relheight=0.12)
        
        self.frame_estilo2 = Frame(self.frame_valores, bd=0, highlightthickness=0,bg=corClara)
        self.frame_estilo2.place(relx=-0.007,rely=0.34,relwidth=1.014,relheight=0.16)
        
        self.frame_estilo3 = Frame(self.frame_valores, bd=0, highlightthickness=0,bg=corEscura)
        self.frame_estilo3.place(relx=-0.007,rely=0.50,relwidth=1.014,relheight=0.16)

        self.frame_estilo4 = Frame(self.frame_valores, bd=0, highlightthickness=0,bg=corClara)
        self.frame_estilo4.place(relx=-0.007,rely=0.66,relwidth=1.014,relheight=0.16)
        
        self.frame_estilo5 = Frame(self.frame_valores, bd=0, highlightthickness=0,bg=corEscura)
        self.frame_estilo5.place(relx=-0.007,rely=0.82,relwidth=1.014,relheight=0.16)
    
    def widgets_frame_valores(self):
        #Label Fixas Odd e Stake
        self.lb_odd = Label(self.frame_estilo1,bg=corEscura,fg=corTitulo, font=('verdana',12, 'bold'), text="Odd")
        self.lb_odd.place(relx=0.33,rely=0.1,relwidth=0.3,relheight=0.8)

        self.lb_stk = Label(self.frame_estilo1,bg=corEscura,fg=corTitulo, font=('verdana',12, 'bold'), text="Stake")
        self.lb_stk.place(relx=0.69,rely=0.1,relwidth=0.3,relheight=0.8)

        #botões
        self.bt_alte = Button(self.frame_estilo5,text="Alterar",bg=corClara,bd=2, command=self.alterar)
        self.bt_alte.place(relx=0.1,rely=0.1,relwidth=0.2,relheight=0.8)
        self.bt_salv = Button(self.frame_estilo5,text="Salvar",bg=corClara,bd=2)
        self.bt_salv.place(relx=0.7,rely=0.1,relwidth=0.2,relheight=0.8)
        self.bt_apagar = Button(self.frame_estilo5,text="Apagar",bg=corClara,bd=2,command=self.apagar_labels)
        self.bt_apagar.place(relx=0.4,rely=0.1,relwidth=0.2,relheight=0.8)
    def widgets_frame_url(self):

        self.lb_url = Label(self.frame_url, bg=corFundoFrame,fg=corTitulo, font=('verdana',12, 'bold'), text="Digite aqui sua URL:")
        self.lb_url.place(relx=0.1,rely=0.1,relwidth=0.8,relheight=0.2)

        self.entry_url = Entry(self.frame_url,bg=corClara,font=('verdana',8, 'bold'),bd=0)
        self.entry_url.place(relx=0.1,rely=0.4,relwidth=0.8,relheight=0.2)

        self.bt_url_comf = Button(self.frame_url, text="Comfirmar", command=self.up_url,bg=corClara,bd=2)
        self.bt_url_comf.place(relx=0.55,rely=0.7,relwidth=0.35,relheight=0.2)

        self.bt_url_apg = Button(self.frame_url, text="Apagar",command= self.apagar,bg=corClara,bd=2)
        self.bt_url_apg.place(relx=0.1,rely=0.7,relwidth=0.35,relheight=0.2)



    #Label Nome dos jogos 
    def label_nome(self,texto):
        self.lb_nome = Label(self.frame_valores, bg=corFundoFrame,fg=corTitulo, font=('verdana',18, 'bold'), text=texto)
        self.lb_nome.place(relx=0.05,rely=0.015,relwidth=0.9,relheight=0.2)

    #Labels e Entry Da Casas 1
    def casa_1(self,nome,odd,stake):
        #Label Nome
        self.lb_nome_c1 = Label(self.frame_estilo2, bg=corClara ,fg=corTitulo, font=('verdana',10, 'bold'),text=nome)
        self.lb_nome_c1.place(relx=0.01,rely=0.2,relwidth=0.3,relheight=0.6)
        #Label Odd
        self.lb_odd_c1 = Label(self.frame_estilo2, bg=corClara ,fg=corFonte, font=('verdana',10, 'bold'), text=odd)
        self.lb_odd_c1.place(relx=0.37,rely=0.2,relwidth=0.12,relheight=0.6)
        #Entry odd
        self.entry_odd_c1 = Entry(self.frame_estilo2,bg=corClara,font=('verdana',8, 'bold'),bd=0.5) 
        self.entry_odd_c1.place(relx=0.5,rely=0.2,relwidth=0.11,relheight=0.6)
        #Label Stake
        self.lb_stake_c1 = Label(self.frame_estilo2, bg=corClara,fg=corFonte, font=('verdana',10, 'bold'), text=stake)
        self.lb_stake_c1.place(relx=0.72,rely=0.2,relwidth=0.12,relheight=0.6)
        #Entry Stake
        self.entry_stk_c1 = Entry(self.frame_estilo2,bg=corClara,font=('verdana',8, 'bold'),bd=0.5) 
        self.entry_stk_c1.place(relx=0.85,rely=0.2,relwidth=0.11,relheight=0.6)

    #Labels e Entry Da Casas 2
    def casa_2(self,nome,odd,stake):
        #Label Nome
        self.lb_nome_c2 = Label(self.frame_estilo3, bg=corEscura ,fg=corTitulo, font=('verdana',10, 'bold'),text=nome)
        self.lb_nome_c2.place(relx=0.01,rely=0.2,relwidth=0.3,relheight=0.6)
        #Label Odd
        self.lb_odd_c2 = Label(self.frame_estilo3, bg=corEscura,fg=corFonte, font=('verdana',10, 'bold'), text=odd)
        self.lb_odd_c2.place(relx=0.37,rely=0.2,relwidth=0.12,relheight=0.6)
        #Entry Odd
        self.entry_odd_c2 = Entry(self.frame_estilo3,bg=corEscura,font=('verdana',8, 'bold'),bd=0.5) 
        self.entry_odd_c2.place(relx=0.5,rely=0.2,relwidth=0.11,relheight=0.6)
        #Label Stake
        self.lb_stake_c2 = Label(self.frame_estilo3, bg=corEscura,fg=corFonte, font=('verdana',10, 'bold'), text=stake)
        self.lb_stake_c2.place(relx=0.72,rely=0.2,relwidth=0.12,relheight=0.6)
        #Entry Stake
        self.entry_stk_c2 = Entry(self.frame_estilo3,bg=corEscura,font=('verdana',8, 'bold'),bd=0.5) 
        self.entry_stk_c2.place(relx=0.85,rely=0.2,relwidth=0.11,relheight=0.6)

    #Labels e Entry Da Casas 3
    def casa_3(self,nome,odd,stake):
        #Label Nome
        self.lb_nome_c3 = Label(self.frame_estilo4, bg=corClara ,fg=corTitulo, font=('verdana',10, 'bold'),text=nome)
        self.lb_nome_c3.place(relx=0.01,rely=0.2,relwidth=0.3,relheight=0.6)
        #Label Odd
        self.lb_odd_c3 = Label(self.frame_estilo4, bg=corClara,fg=corFonte, font=('verdana',10, 'bold'), text=odd)
        self.lb_odd_c3.place(relx=0.37,rely=0.2,relwidth=0.12,relheight=0.6)
        #Entry Odd
        self.entry_odd_c3 = Entry(self.frame_estilo4,bg=corClara,font=('verdana',8, 'bold'),bd=0.5) 
        self.entry_odd_c3.place(relx=0.5,rely=0.2,relwidth=0.11,relheight=0.6)
        #Label Stake
        self.lb_stake_c3 = Label(self.frame_estilo4, bg=corClara,fg=corFonte, font=('verdana',10, 'bold'), text=stake)
        self.lb_stake_c3.place(relx=0.72,rely=0.2,relwidth=0.12,relheight=0.6)
        #Entry Stake
        self.entry_stk_c3 = Entry(self.frame_estilo4,bg=corClara,font=('verdana',8, 'bold'),bd=0.5)
        self.entry_stk_c3.place(relx=0.85,rely=0.2,relwidth=0.11,relheight=0.6)

Application()