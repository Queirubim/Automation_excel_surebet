from bs4 import BeautifulSoup
import openpyxl
import requests
from datetime import date,datetime


url = requests.get("https://pt.surebet.com/calculator/show/coYLe3KEojdyhKIw").content

soup =BeautifulSoup(url, 'html.parser')
soup.encode= "utf-8"


# Coleta de dados
dataAtual = date.today().strftime('%d/%m/%Y')
horaAtual = datetime.now().strftime('%H:%M')

jogo = soup.find("span", attrs={"class": "calc_event"}).text

infoC1 = soup.find("tr", attrs={"data-number": "0"})
nomeC1 = infoC1.find("td", attrs={"class": "booker_c"}).text
oddC1 = infoC1.find("input", attrs={"class": "koefficient form-control inline-input w-number"})
entradaC1 = infoC1.find("td", attrs= {"coeff_c coeff_name"}).text

infoC2 = soup.find("tr", attrs={"data-number": "1"})
nomeC2 = infoC2.find("td", attrs={"class": "booker_c"}).text
oddC2 = infoC2.find("input", attrs={"class": "koefficient form-control inline-input w-number"})
entradaC2 = infoC2.find("td", attrs= {"coeff_c coeff_name"}).text

infoC3 = soup.find("tr", attrs={"data-number": "2"})
if infoC3 != None:
    nomeC3 = infoC3.find("td", attrs={"class": "booker_c"}).text
    oddC3 = infoC3.find("input", attrs={"class": "koefficient form-control inline-input w-number"})
    entradaC3 = infoC3.find("td", attrs= {"coeff_c coeff_name"}).text
    nomeC3.encode("utf8")


nomeC1.encode("utf8")
nomeC2.encode("utf8")



#planilha
caminho = "SureBet3.xlsx"
planilha = openpyxl.load_workbook(caminho)
pagina = planilha['Planilha de Apostas']

#acha a proxima linha vazia
vazia = 17
procurando = True
while procurando:
    for row in pagina.iter_rows(min_row=vazia):
        print(vazia)
        if(row[3].value == None):
            procurando = False
        else:
            vazia += 3
        break
    

print(f'{vazia} terminou')
#Joga dados na planilha
pagina.cell(row=vazia, column=3).value = dataAtual
pagina.cell(row=vazia, column=4).value = nomeC1
pagina.cell(row=vazia+1, column=4).value = nomeC2
pagina.cell(row=vazia, column=5).value = horaAtual
pagina.cell(row=vazia, column=6).value = jogo
pagina.cell(row=vazia, column=7).value = entradaC1
pagina.cell(row=vazia+1, column=7).value = entradaC2
pagina.cell(row=vazia, column=8).value = 2.000
pagina.cell(row=vazia+1, column=8).value = 2.000
pagina.cell(row=vazia+2, column=8).value = 2.000
pagina.cell(row=vazia, column=9).value = float(oddC1['value'])
pagina.cell(row=vazia+1, column=9).value = float(oddC2['value'])
if infoC3 != None:
    pagina.cell(row=vazia+2, column=9).value = float(oddC3['value'])
    pagina.cell(row=vazia+2, column=4).value = nomeC3
    pagina.cell(row=vazia+2, column=7).value = entradaC3

planilha.save(caminho)